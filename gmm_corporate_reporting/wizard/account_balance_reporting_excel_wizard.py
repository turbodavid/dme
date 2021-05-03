
# -*- encoding: utf-8 -*-
###########################################################################
#    Module Writen to OpenERP, Open Source Management Solution
#
#    Copyright (c) 2019 Grupo MORSA - http://www.morsa.com.mx
#    All Rights Reserved.
############################################################################
#    Coded by:
#       David Alberto Perez Payán (davidperez@dmesoluciones.com)
############################################################################
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

#from openerp.addons.web.controllers.main import Session
#import datetime as dt
import locale
from datetime import datetime, timedelta
from openerp.exceptions import Warning as UserError
from openerp import fields, models, api
import openpyxl as pyxl
from openpyxl.styles import NamedStyle, Font, Color, Alignment, Border, Side, colors
from openpyxl.utils import get_column_letter
import cStringIO
import StringIO
import base64
import tempfile
import binascii
from copy import copy


EXPENSES_SHEET = 'Gastos GMM'
MONTHNAMES = {'01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO', '04': 'ABRIL',
              '05': 'MAYO', '06': 'JUNIO', '07': 'JULIO', '08': 'AGOSTO',
              '09': 'SEPTIEMBRE', '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
              }
ZONAS = [
    ['', 'TOTAL GRUPO'],
    ['occidente', 'OCCIDENTE'],
    ['noroeste', 'NOROESTE'],
    ['diesel', 'División Diesel'],
    ['sureste', 'SURESTE'],
    ['cslthsur', 'CS LTH Sureste'],
    ['cslthnor', 'CS LTH Noroeste/Dgo.']
]
class AccountBalanceReportingExcelWizard(models.TransientModel):
    _name = 'account.balance.reporting.excel.wizard'
    _description = 'Wizard for Excel Corporate Reports Generation'


    def _get_active_report(self):
        active_id = self.env.context.get('active_id') or 0
        return active_id

    any_report_id = fields.Many2one(
        'account.balance.reporting',
        'Report to fill in Excel',
        default=lambda self: self._get_active_report(),
        help='Select the Report for Excel generation'
        )

    balance_report_id = fields.Many2one(
        'account.balance.reporting',
        'Report for General Balance',
        help='Select the General Balance Report for Excel generation'
        )

    profitloss_report_id = fields.Many2one(
        'account.balance.reporting',
        'Report for Profit/Loss',
        help='Select the Profit/Loss Report for Excel generation'
     )

    profitloss_by_branch_report_id = fields.Many2one(
        'account.balance.reporting',
        'Report for Profit/Loss by branch',
        help='Select the Profit/Loss by branch Report for Excel generation'
     )

    expenses_report_id = fields.Many2one(
        'account.balance.reporting',
        'Report for Detailed Expenses',
        help='Select the Detailed Expenses Report for Excel generation'
     )

    corporate_report_id = fields.Many2one(
        'account.balance.reporting',
        'Report for Corporate Expenses',
        help='Select the Corporate Expenses Report for Excel generation'
     )

    upto_current_period = fields.Many2one(
        'account.period',
        'Current',
        help='Select up to current period to process'
    )

    upto_previous_period = fields.Many2one(
        'account.period',
        'Previous',
        help='Select up to previous period to process'
    )

    report = fields.Binary(string='Excel Template File', help='Select the Excel File to fill')
    filename = fields.Char(string='File Name')
    wb = ''
    curr_zone_expenses = ''
    prev_zone_expenses = ''
    curr_zone_sales = ''
    prev_zone_sales = ''
    sales_report_line = ''

    curr_noroeste_expenses = ''
    prev_noroeste_expenses = ''
    curr_noroeste_sales = ''
    prev_noroeste_sales = ''


    @api.multi
    @api.onchange('any_report_id', 'balance_report_id')
    def onchange_any_report_id(self):

        if self._context.get('reportmode') != 'single':
            report_obj = self.balance_report_id
        else:
            report_obj = self.any_report_id

        curr_periods = report_obj.current_period_ids.ids
        prev_periods = report_obj.previous_period_ids.ids

        domain = {'upto_current_period': [('id', 'in', curr_periods)],
                  'upto_previous_period': [('id', 'in', prev_periods)]
                  }

        if curr_periods:
            self.upto_current_period = curr_periods[len(curr_periods)-1]

        if prev_periods:
            self.upto_previous_period = prev_periods[len(prev_periods)-1]

        return {'domain': domain}

    @api.multi
    def do_excel_report(self):

        locale.setlocale(locale.LC_TIME, "es_MX.UTF-8")
        fp2 = tempfile.NamedTemporaryFile(suffix=".xlsx")
        fp2.write(binascii.a2b_base64(self.report))  # self.xls_file is your binary field
        fp2.seek(0)
        self.wb = pyxl.load_workbook(fp2.name)
        try:
            if self._context.get('reportmode') == 'single':
                if self.any_report_id.name[:10] == "Gastos GMM":
                    self._gastos_gmm()
                else:
                    self._process_report(self.any_report_id, self.any_report_id.template_id.use_materialized_view)
            else:
                self._process_report(self.profitloss_report_id)
                self._process_report(self.balance_report_id)
                if self.profitloss_by_branch_report_id:
                    self._process_report(self.profitloss_by_branch_report_id,
                                         self.profitloss_by_branch_report_id.template_id.use_materialized_view)
                if self.expenses_report_id:
                    self._gastos_gmm()
                if self.corporate_report_id:
                    self._process_report(self.corporate_report_id,
                                         self.corporate_report_id.template_id.use_materialized_view)
            try:
                fp = cStringIO.StringIO()
                print fp
                self.wb.save(fp)
                out = base64.encodestring(fp.getvalue())
                self.write({'report': out, 'filename': 'inf_corporativo.xlsx'})
            except Exception as e:
                raise e
            finally:
                fp.close()
        except Exception as e:
            raise e
        finally:
            fp2.close()
            self.wb.close()

        # return {
        #     'type': 'ir.actions.act_url',
        #     'url': '/web/binary/saveas?model=account.balance.reporting.excel.wizard&field=report&id=%s&filename=inf_corporativo.xlsx&filename_field=inf_corporativo.xlsx' % (
        #         self.id),
        #     'target': 'self',
        # }

        return {
           'type': 'ir.actions.act_url',
           'url': '/web/binary/download_document?model=account.balance.reporting.excel.wizard&field=report&id=%s&filename=inf_corporativo.xlsx' % (
               self.id),
           'target': 'self',
        }

    def _process_report(self, reportobj, useexpenses=False):

        rlines = reportobj.line_ids
        balance_mode = 'debit - credit'
        if int(reportobj.template_id.balance_mode) > 1:
            balance_mode = 'credit - debit'

        monthcurr = max(reportobj.current_period_ids.ids)
        monthprev = max(reportobj.previous_period_ids.ids)

        currperiod = None
        prevperiod = None
        calccurrperiod = (monthcurr > self.upto_current_period.id)
        calcprevperiod = (monthprev > self.upto_previous_period.id)

        if calccurrperiod:
            currperiod = self.upto_current_period.id
            monthcurr = currperiod

        if calcprevperiod:
            prevperiod = self.upto_previous_period.id
            monthprev = prevperiod

        for rl in rlines:

            tmpl = rl.template_line_id
            exlines = tmpl.template_line_excel_ids
            amtacumcurr = rl.current_value
            amtacumprev = rl.previous_value
            forcecalc = calccurrperiod or calcprevperiod or (len(exlines) > 1)

            for exl in exlines:

                if not (exl.cell_acum_current or exl.cell_acum_previous):
                    continue
                print 'Excel Name, Id, Codigo: %s, %sm %s' % (tmpl.name, exl.id, tmpl.code),

                cellacumcurr = exl.cell_acum_current
                cellacumprev = exl.cell_acum_previous
                cellmonthprev = exl.cell_month_previous
                cellmonthcurr = exl.cell_month_current
                amtmonthcurr = 0.0
                amtmonthprev = 0.0

                if forcecalc or len(exl.operating_unit_ids) > 0:
                    amtacumcurr, amtacumprev = self._get_balance_vm(rl, exl, False) if useexpenses \
                                                        else self._get_balance_ous(rl, exl, False)

                if (cellmonthcurr or cellmonthprev):
                    amtmonthcurr, amtmonthprev = self._get_balance_vm(rl, exl, True) if useexpenses \
                                                        else self._get_balance_ous(rl, exl, True)


                sheetacum = self.wb.get_sheet_by_name(exl.sheet_acum)
                sheetmonth = self.wb.get_sheet_by_name(exl.sheet_month)

                if cellacumcurr:
                    sheetacum[cellacumcurr] = amtacumcurr

                if cellacumprev:
                    sheetacum[cellacumprev] = amtacumprev

                if cellmonthcurr:
                    sheetmonth[cellmonthcurr] = amtmonthcurr

                if cellmonthprev:
                    sheetmonth[cellmonthprev] = amtmonthprev

                #except Exception as x:
                #    print repr(x)
                #    pass

        return True

    def _get_balance_ous(self, rline, exline, month=False):

        template_line_id = rline.template_line_id
        sign = -1 if template_line_id.negate else 1

        res = {'current': {'amount': 0.0,
                           'table': 'account_balance_reporting_line_account_move_line_rel abrlmc ',
                           'report_id': 'abrlmc.account_balance_reporting_line_id ',
                           'field_rel': 'abrlmc.account_move_line_id',
                           'calcexpr': template_line_id.current_value,
                           'period_id': self.upto_current_period.id,
                           'sql': ''
                           },
               'previous': {'amount': 0.0,
                            'table': 'account_balance_reporting_line_previous_move_line_rel abrlmp ',
                            'report_id': 'abrlmp.report_id',
                            'field_rel': 'abrlmp.line_id',
                            'calcexpr': template_line_id.previous_value,
                            'period_id': self.upto_previous_period.id,
                            'sql': ''
                            }
               }

        join = ''
        if len(exline.operating_unit_ids) > 0:
            join = " inner join account_balance_reporting_template_line_excel_ou abrtlexou on " \
                                "(aml.operating_unit_id = abrtlexou.operating_unit_id and abrtlexou.excel_id = %s)" % exline.id
        print " Join: %s" % join
        for k in res.keys():
            sqlwhere = " %s = %s and aml.period_id %s %s " \
                            % (res[k]['report_id'], rline.id, '=' if month else '<=', res[k]['period_id'])
            sqlwhere += self._get_negative_accounts(res[k]['calcexpr'])

            sql = """
                    select '%s' fiscalyear, coalesce( sum(debit-credit), 0.00) balance                   
                    from %s inner join account_move_line aml on (aml.id = %s) %s
                    where %s 
                """ % (k, res[k]['table'], res[k]['field_rel'], join, sqlwhere)
            res[k]['sql'] = sql

        sql = res['current']['sql'] + " union " + res['previous']['sql'] + " group by 1;"

        self.env.cr.execute(sql)
        resbalance = self.env.cr.dictfetchall()
        for balance in resbalance:
            res[balance.get('fiscalyear')]['amount'] = balance.get('balance') * sign or 0.00

        return res['current']['amount'], res['previous']['amount']

    def _get_negative_accounts(self, expr):

        #busca por negativos
        where = ''
        accounts = expr.split(",")
        accids_negative = []
        for acc in accounts:
            if acc[:1] != '-':
                continue

            acc_code = acc[1:]
            if acc_code.startswith('(') and acc_code.endswith(')'):
                acc_code = acc_code[1:-1]

            accobj = self.env['account.account']
            accparent = accobj.search([('code', '=', acc_code)])
            acchilds = accobj.search([('id', 'child_of', accparent.id)])
            accids_negative += acchilds.ids or accparent.ids

        if accids_negative:
            #domain.append(('account_id', 'not in', accids_negative))
            where += " and aml.account_id not in %s" % str(tuple(accids_negative)).replace(",)", ")")

        return where

    def _get_balance_vm(self, rline, exline, month=False):

        template_line_id = rline.template_line_id
        sign = -1 if template_line_id.negate else 1

        maxcurrperiod = self.upto_current_period.id
        mincurrperiod = min(rline.report_id.current_period_ids.ids)
        maxprevperiod = self.upto_previous_period.id
        minprevperiod = min(rline.report_id.previous_period_ids.ids)
        ous = [x.operating_unit_id.id for x in exline.operating_unit_ids]
        sou = (" and g.operating_unit_id in %s " % str(tuple(ous)).replace(",)", ")")) if ous else ''

        res = {'current': {'amount': 0.0,
                           'calcexpr': template_line_id.current_value,
                           'periods': " and g.period_id = %s" % maxcurrperiod if month
                                        else (" and g.period_id between %s and %s" % (mincurrperiod, maxcurrperiod) ),
                           'sql': ''
                           },
               'previous': {'amount': 0.0,
                            'calcexpr': template_line_id.previous_value,
                            'periods':  " and g.period_id = %s" % maxprevperiod if month
                                         else (" and g.period_id between %s and %s" % (minprevperiod, maxprevperiod)),
                            'sql': ''
                            }
               }

        for k in res.keys():
            sqlwhere = " abrl.id = %s and g.account_id not in (select unnest(get_abrl_acc_negatives('%s'))) " \
                       " %s %s " % (rline.id, res[k]['calcexpr'], res[k]['periods'], sou)
            sql = """
                    select '%s' fiscalyear, coalesce( sum(saldo), 0.00) balance                   
                    from gastos_gmm_acc g inner join account_balance_reporting_line abrl on (abrl.id = g.id)
                    where %s 
                """ % (k, sqlwhere)
            res[k]['sql'] = sql

        sql = res['current']['sql'] + " union " + res['previous']['sql'] + " group by 1;"

        self.env.cr.execute(sql)
        resbalance = self.env.cr.dictfetchall()
        for balance in resbalance:
            res[balance.get('fiscalyear')]['amount'] = balance.get('balance') * sign or 0.00

        return res['current']['amount'], res['previous']['amount']

    def _gastos_gmm(self):

        reportobj = self.expenses_report_id or self.any_report_id
        rlines = reportobj.line_ids.sorted(key=lambda r: r.code)
        self.sales_report_line = rlines.filtered(lambda r: r.code == '999')

        sheet = self.wb.get_sheet_by_name('Gastos GMM')
        self._add_sytles()

        #definición de columnas
        lastintcol = reportobj.current_period_ids.ids.index(self.upto_current_period.id) + 2
        #lastintcol = len(reportobj.current_period_ids.ids.index(self.upto_current_period.id)) + 1
        lastcol = get_column_letter(lastintcol)
        rescol = {'monthprevcol': get_column_letter(lastintcol + 2),
                  'difamtmonthcol': get_column_letter(lastintcol + 3),
                  'difpctmonthcol':  get_column_letter(lastintcol + 4),
                  'acumprevcol': get_column_letter(lastintcol + 6),
                  'acumcurrcol': get_column_letter(lastintcol + 7),
                  'difamtacumcol': get_column_letter(lastintcol + 8),
                  'difpctacumcol': get_column_letter(lastintcol + 9)
                  }

        row = 3
        #definir estilos
        for zona in ZONAS:

            row = self._first_header(sheet, row, zona[1], lastintcol, rescol, zona[0] == '')
            if zona[0] == '':
                row = self._excel_line(sheet, rlines, row, zona[0], zona[0], lastintcol, rescol, zona[1], 1)
                continue

            row = self._excel_line(sheet, rlines, row, zona[0], zona[0], lastintcol, rescol, zona[1], 3)
            ous = self.env['operating.unit'].search([('corporate_ou', '=', False),
                                                     ('zone', '=', zona[0]),
                                                     ('code', 'not ilike', '-')], order='code')
            if ous:
                for ou in ous:
                    sql = """
                        select count(*) counter 
                        from gastos_gmm_acc g inner join account_balance_reporting_line abrl on (g.id = abrl.id)
                                    inner join operating_unit ou on (ou.id = g.operating_unit_id) 
                        where (operating_unit_id = %s or ou.code like '%s')
                            and abrl.report_id = %s and period_id between %s and %s;
                    """ % (ou.id, ou.code + '-%', reportobj.id, min(reportobj.current_period_ids.ids), self.upto_current_period.id)
                    self.env.cr.execute(sql)
                    counter = self.env.cr.dictfetchone()
                    if counter['counter'] > 0:
                        row = self._first_header(sheet, row, ou.name, lastintcol, rescol, False)
                        row = self._excel_line(sheet, rlines, row, zona[0], ou.code, lastintcol, rescol, ou.name)

        sheet.freeze_panes = sheet["A9"]

        #es necesario remarcar y hacer merge
        #cuando las columnas sean mayor a uno
        #insertar columnas con sheet.insert_cols(idx=2, [amoount=numcols cuando es mas de una columna])
        #hacer merge a las las columnas
        return True

    def _excel_line(self, sheet, rlines, row, zona, sucursal, lastintcol, rescol, totaltitle, corporatetype=0):

        rowini = row
        lastcol = get_column_letter(lastintcol)
        lfirsttotal = True
        for rline in rlines:

            linecode = int(rline.code[:3]) #codigo del reporte
            linename = rline.name
            print 'LineName:', linename
            currvalue = [0.0]
            prevalue = [0.0]

            sumacumcurr = 0.0
            sumacumprev = 0.0
            if linecode == 999:
                continue

            print zona, sucursal, rline
            if linecode < 636:
                currvalue, prevalue = self._get_expenses_balances(rline, zona if zona == sucursal else '', sucursal)
            else:
                if lfirsttotal:
                    row = self._totaliza(sheet, rowini, row, lastintcol, rescol, "GASTOS:")
                    rowini = row - 2
                    lfirsttotal = False

                if corporatetype == 1:
                    if linecode == 641:
                        currvalue, prevalue = self._get_expenses_balances(rline, zona, sucursal, 1)
                        linename = "DIRECCION GENERAL"
                    elif linecode == 646:
                        currvalue, prevalue = self._get_expenses_balances(rline, zona, sucursal, 2)
                        linename = "CORPORATIVO NACIONAL"
                    else:
                        currvalue, prevalue = self._get_expenses_balances(rline, zona, sucursal, 0)
                elif corporatetype == 3:
                    if linecode == 641:
                        continue
                    elif linecode == 646:
                        linename = "CORPORATIVO %s" % totaltitle
                        #hay que salvar noroeste porque se utiliza para DIESEL
                        if zona == 'noroeste':
                            currvalue, prevalue = self._get_expenses_balances(rline, zona, sucursal, 5)
                            self.curr_zone_sales, self.prev_zone_sales = \
                                   self._get_expenses_balances(self.sales_report_line, zona, sucursal, 5)
                            self.curr_noroeste_expenses = currvalue
                            self.prev_noroeste_expenses = prevalue
                            self.curr_noroeste_sales = self.curr_zone_sales
                            self.prev_noroeste_sales = self.prev_zone_sales
                        elif zona == 'diesel':
                            self.curr_zone_expenses = self.curr_noroeste_expenses
                            self.prev_zone_expenses = self.prev_noroeste_expenses
                            self.curr_zone_sales = self.curr_noroeste_sales
                            self.prev_zone_sales = self.prev_noroeste_expenses
                            continue
                        else:
                            currvalue, prevalue = self._get_expenses_balances(rline, zona, sucursal, 3)
                            self.curr_zone_sales, self.prev_zone_sales = \
                                self._get_expenses_balances(self.sales_report_line, zona, sucursal, 4)
                        self.curr_zone_expenses = currvalue
                        self.prev_zone_expenses = prevalue
                    else:
                        currvalue, prevalue = self._get_expenses_balances(rline, zona, sucursal, 0)
                else:
                    if linecode == 641:
                        continue
                    elif linecode == 646:
                        xsucname = filter(lambda x: x[0] == zona, ZONAS)[0][1]
                        if xsucname == 'División Diesel':
                            xsucname = 'NOROESTE'
                        linename = "CORPORATIVO %s" % xsucname
                        curr_sales, prev_sales = self._get_expenses_balances(self.sales_report_line, '', sucursal)
                        # para el total acumulado, hay que aplicar el prorrateo de otra manera

                        curr_zone_sales = sum(self.curr_zone_sales) if sum(self.curr_zone_sales) != 0 else 1
                        prev_zone_sales = sum(self.prev_zone_sales) if sum(self.prev_zone_sales) != 0 else 1
                        sumacumcurr = sum(self.curr_zone_expenses) * (sum(curr_sales) / curr_zone_sales)
                        sumacumprev = sum(self.prev_zone_expenses) * (sum(prev_sales) / prev_zone_sales)
                        currvalue, prevalue = self._get_prorrateos(curr_sales, prev_sales)
                    else:
                        currvalue, prevalue = self._get_expenses_balances(rline, '', sucursal)

            row += self._fill_excel_line(sheet, linename, row, currvalue, prevalue, lastintcol, rescol, sumacumcurr, sumacumprev)

        self._totaliza(sheet, rowini, row, lastintcol, rescol, totaltitle)

        return row + 2

    def _get_prorrateos(self, curr_sales, prev_sales):
        curr_value = curr_sales
        prev_value = prev_sales

        for i in range(0, len(curr_sales)):
            curr_value[i] = self.curr_zone_expenses[i] * (curr_sales[i] / self.curr_zone_sales[i] if self.curr_zone_sales[i] != 0 else 1)
            prev_value[i] = self.prev_zone_expenses[i] * (prev_sales[i] / self.prev_zone_sales[i] if self.prev_zone_sales[i] != 0 else 1)

        return curr_value, prev_value

    def _fill_excel_line(self, sheet, linename, row, currvalue, prevalue, lastintcol, rescol, sumcurrvalue=0.0, sumprevalue=0.0):

        if sum(currvalue) == 0.0 and sum(prevalue) == 0.0 \
                and sumcurrvalue == 0.0 and sumprevalue == 0.0:
            return 0

        lastcol = get_column_letter(lastintcol)
        srow = str(row)
        col = 1
        self._set_cell_value(sheet.cell(row=row, column=col), linename, "number_style")
        for icol in range(2, lastintcol + 1):
            col = icol
            self._set_cell_value(sheet.cell(row=row, column=col), currvalue[col - 2] or 0, "number_style")

        # para el mes
        cellcurr = lastcol + srow
        cellprev = rescol['monthprevcol'] + srow
        celldif = rescol['difamtmonthcol'] + srow
        self._set_cell_value(sheet[cellprev], prevalue[len(prevalue) - 1], "number_style")
        self._insert_formula(sheet[celldif], '', cellcurr, cellprev)
        self._insert_formula(sheet[rescol['difpctmonthcol'] + srow], celldif, cellcurr, cellprev, "%")

        # para el año
        cellcurr = rescol['acumcurrcol'] + srow
        cellprev = rescol['acumprevcol'] + srow
        celldif = rescol['difamtacumcol'] + srow
        self._set_cell_value(sheet[cellcurr], sum(currvalue) if sumcurrvalue == 0.0 else sumcurrvalue, "number_style")
        self._set_cell_value(sheet[cellprev], sum(prevalue) if sumprevalue == 0.0 else sumprevalue, "number_style")
        self._insert_formula(sheet[celldif], '', cellcurr, cellprev)
        self._insert_formula(sheet[rescol['difpctacumcol'] + srow], celldif, cellcurr, cellprev, "%")

        return 1

    def _totaliza(self, sheet, rowini, row, lastintcol, rescol, totaltitle):

        srow = str(row - 1)
        srowini = str(rowini)
        lastcol = get_column_letter(lastintcol)

        self._set_cell_value(sheet.cell(row=row, column=1), "TOTAL %s: " % totaltitle)
        for icol in range(2, lastintcol + 1):
            col = icol
            scol = get_column_letter(col)
            self._insert_formula(sheet.cell(row=row, column=col), '', scol + srowini, scol + srow, ":", True)

        for k in rescol.keys():
            if k in ['difpctacumcol', 'difpctacumcol']:
                continue
            scol = rescol[k]
            self._insert_formula(sheet[scol + str(row)], '', scol + srowini, scol + srow, ":", True)

        srow = str(row)
        self._insert_formula(sheet[rescol['difpctacumcol'] + str(row)],
                             rescol['difamtacumcol'] + srow,
                             rescol['acumcurrcol'] + srow,
                             rescol['acumprevcol'] + srow, "%")

        self._insert_formula(sheet[rescol['difpctmonthcol'] + str(row)],
                             rescol['difamtmonthcol'] + srow,
                             lastcol + srow,
                             rescol['acumprevcol'] + srow, "%")

        return row + 2

    def _insert_formula(self, celldif, cellamtdif, cellcurr, cellprev, type='-', bold=False):
        formula = '=+%s%s%s' % (cellcurr, type, cellprev)
        style = "number_style"
        if type == '%':
            formula = '=IF(%s = 0.0, IF(%s = 0.0, 0.0, 1 ), IF( %s <> 0.0, +%s/%s, -1))' \
                      % (cellprev, cellcurr, cellcurr, cellamtdif, cellcurr)
            style = "percent_style"
        elif type == ':':
            formula = '=sum(%s)' % formula[2:]

        style += "_bold" if bold else ''
        self._set_cell_value(celldif, formula, style)

    def _get_expenses_balances(self, rline, zona, sucursal, corporatetype=0):
        #corporatetype puede tomar los siguientes valores

        rescorporate = {1: " and ou.code = '501' ", #DIRECCION GENERAL
                        2: """ and (ou.corporate_ou and (ou."zone" is null or ou."zone"='') and ou.code != '501') """, #CORPORATIVO NACIONAL
                        3: """ and (ou.corporate_ou and ou."zone" = '%s') """ % sucursal, #CORPORATIVO REGIONAL
                        4: """ and (ou."zone" = '%s') """ % sucursal, #PARA LAS VENTAS REGIONALES
                        5: """ and (ou."zone" in ('noroeste','diesel')) """ #PARA VENTAS Y GASTOS NOROESTE
                    }

        template_line_id = rline.template_line_id
        sign = -1 if (template_line_id.negate or corporatetype == 2) else 1

        maxcurrperiod = self.upto_current_period.id
        mincurrperiod = min(rline.report_id.current_period_ids.ids)
        lencurrperiod = rline.report_id.current_period_ids.ids.index(self.upto_current_period.id) + 1
        #lencurrperiod = len(rline.report_id.current_period_ids.ids)
        maxprevperiod = self.upto_previous_period.id
        minprevperiod = min(rline.report_id.previous_period_ids.ids)
        lenprevperiod = rline.report_id.previous_period_ids.ids.index(self.upto_previous_period.id) + 1
        #lenprevperiod = len(rline.report_id.previous_period_ids.ids)

        sou = ''
        if sucursal or corporatetype:
            if corporatetype:
                sou = rescorporate[corporatetype]
            else:
                sou = (" and ou.zone = '%s'" % sucursal) if zona else (" and ou.code like '%s'" % (sucursal + '%'))

        res = {'current': {'amount': [0.00 for x in range(0, lencurrperiod)],
                       'calcexpr': template_line_id.current_value,
                       'periods': "  and g.period_id between %s and %s" % (mincurrperiod, maxcurrperiod),
                       'sql': ''
                       },
           'previous': {'amount': [0.00 for x in range(0, lenprevperiod)],
                        'calcexpr': template_line_id.previous_value,
                        'periods': " and g.period_id between %s and %s" % (minprevperiod, maxprevperiod),
                        'sql': ''
                        }
           }

        for k in res.keys():
            sqlwhere = " abrl.id = %s " \
                       "and g.account_id not in (select unnest(get_abrl_acc_negatives('%s'))) " \
                       " %s %s " % (rline.id, res[k]['calcexpr'], res[k]['periods'], sou)
            sql = """
                select abrl.code,  '%s' fiscalyear,  left(ap.code, 2) period_code, sum(saldo) balance
                from gastos_gmm_acc g 
                    inner join account_balance_reporting_line as abrl on (g.id = abrl.id)
                    inner join account_period ap on (ap.id = g.period_id)
                    inner join operating_unit ou on (ou.id = g.operating_unit_id)
                where %s
                group by 1, 2, 3
                """ % (k, sqlwhere)
            res[k]['sql'] = sql

        sql = res['current']['sql'] + " union " + res['previous']['sql'] + " order by 1,2,3;"

        self.env.cr.execute(sql)
        resbalance = self.env.cr.dictfetchall()
        for balance in resbalance:
            res[balance.get('fiscalyear')]['amount'][int(balance.get('period_code'))-1] = balance.get('balance') * sign or 0.00

        return res['current']['amount'], res['previous']['amount']

    def _first_header(self, ws, row, title, lastintcol, rescol, lfirst=False):
        # define estilo de la que será base

        smonth = datetime.strftime(datetime.strptime(self.upto_current_period.date_stop, '%Y-%m-%d'),
                                   '%B de %Y').upper()
        syearprev = self.upto_previous_period.code[-4:]

        self._set_cell_value(ws.cell(row=row, column=1), self.env.user.company_id.name)
        self._set_cell_value(ws.cell(row=row+1, column=1), 'RELACIÓN ANALITICA DE GASTOS')
        self._set_cell_value(ws.cell(row=row+2, column=1),  smonth)

        row += 4
        srowfirst = str(row)
        row += 1
        srow = str(row)
        if lfirst:
            col_month_width = ws.column_dimensions["B"].width
            self._set_cell_value(ws["B8"], MONTHNAMES['01'], "month_style")
            for col in range(3, lastintcol + 1):
                target_cell = get_column_letter(col)
                ws.insert_cols(col)
                ws.column_dimensions[target_cell].width = col_month_width
                target_cell += srow
                self._set_cell_value(ws[target_cell], MONTHNAMES[str(col - 1).zfill(2)], "month_style")

            for k in rescol.keys():
                ws.column_dimensions[rescol[k]].width = col_month_width

            col_month_width = ws.column_dimensions["AZ"].width
            ws.column_dimensions[get_column_letter(lastintcol + 1)].width = col_month_width
            ws.column_dimensions[get_column_letter(lastintcol + 5)].width = col_month_width

        else:
            for col in range(2, lastintcol + 1):
                self._set_cell_value(ws.cell(row=row, column=col), MONTHNAMES[str(col - 1).zfill(2)], "month_style")
            ws.row_dimensions[row-1].height = ws.row_dimensions[7].height

        self._set_cell_value(ws.cell(row=row-1, column=2), "GASTOS MENSUALES %s" % smonth[-4:], "month_expenses_style")
        srange = "%s:%s" % ('B' + srowfirst, get_column_letter(lastintcol) + srowfirst)
        ws.merge_cells(srange)

        self._set_cell_value(ws[rescol['monthprevcol'] + srowfirst], "MENSUAL " + smonth[:len(smonth)-8], "month_expenses_style")
        self._set_cell_value(ws[rescol['monthprevcol'] + srow], syearprev, "month_style")

        self._set_cell_value(ws[rescol['difamtmonthcol'] + srowfirst], 'DIF. vs mismo mes %s' % syearprev, "month_expenses_style")
        self._set_cell_value(ws[rescol['difamtmonthcol'] + srow], '$ $ $', "month_style")
        self._set_cell_value(ws[rescol['difpctmonthcol'] + srow], '% % %', "month_style")
        srange = "%s:%s" % (rescol['difamtmonthcol'] + srowfirst, rescol['difpctmonthcol'] + srowfirst)
        ws.merge_cells(srange)

        self._set_cell_value(ws[rescol['acumprevcol'] + srowfirst], 'ACUMULADO', "month_expenses_style")
        self._set_cell_value(ws[rescol['acumprevcol'] + srow], syearprev, "month_style")
        self._set_cell_value(ws[rescol['acumcurrcol'] + srow], smonth[-4:], "month_style")
        srange = "%s:%s" % (rescol['acumprevcol'] + srowfirst, rescol['acumcurrcol'] + srowfirst)
        ws.merge_cells(srange)

        self._set_cell_value(ws[rescol['difamtacumcol'] + srowfirst], 'DIF. vs PERIODO ANT', "month_expenses_style")
        self._set_cell_value(ws[rescol['difamtacumcol'] + srow], '$ $ $', "month_style")
        self._set_cell_value(ws[rescol['difpctacumcol'] + srow], '% % %', "month_style")
        srange = "%s:%s" % (rescol['difamtacumcol'] + srowfirst, rescol['difpctacumcol'] + srowfirst)
        ws.merge_cells(srange)


        row += 1
        self._set_cell_value(ws.cell(row=row, column=1), title)
        row += 1

        return row

    def _set_cell_value(self, cell, value=None, style="titles_style"):

        if value or value == 0.0:
            cell.value = value

        if style:
            cell.style = style

    def _add_sytles(self):

        border_side = Side(border_style='medium')
        border_month = Border(left=border_side, right=border_side, bottom=border_side)
        border_heads = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        font_month_expenses = Font(name='Calibri', size=11, bold=True, color='FF3C3C3C')
        font_titles = Font(name='Calibri', size=11, bold=True)
        font_normal = Font(name='Calibri', size=11)
        aligment = Alignment(vertical='center', horizontal='center', wrapText=True)

        if self._not_in_style("titles_style"):
            titles_sytle = NamedStyle(name="titles_style", font=font_titles)
            self.wb.add_named_style(titles_sytle)

        if self._not_in_style("month_expenses_style"):
            month_expenses_style = NamedStyle(name="month_expenses_style", font=font_month_expenses,
                                              border=border_heads, alignment=aligment)
            self.wb.add_named_style(month_expenses_style)

        if self._not_in_style("month_style"):
            month_style = NamedStyle(name="month_style", font=font_titles, border=border_month, alignment=aligment)
            self.wb.add_named_style(month_style)

        if self._not_in_style("number_style"):
            number_style = NamedStyle(name="number_style", font=font_normal, number_format="#,###.00;[RED]-#,###.00")
            self.wb.add_named_style(number_style)

        if self._not_in_style("percent_style"):
            percent_style = NamedStyle(name="percent_style", font=font_normal, number_format="0.00%;[RED]-0.00%")
            self.wb.add_named_style(percent_style)

        if self._not_in_style("number_style_bold"):
            number_style_bold = NamedStyle(name="number_style_bold", font=font_titles, number_format="#,###.00;[RED]-#,###.00")
            self.wb.add_named_style(number_style_bold)

        if self._not_in_style("percent_style_bold"):
            percent_style_bold = NamedStyle(name="percent_style_bold", font=font_titles, number_format="0.00%;[RED]-0.00%")
            self.wb.add_named_style(percent_style_bold)

    def _not_in_style(self, style):
        return style not in self.wb.named_styles

    def _copy_cell(self, ws, source_cell, target_cell):

        ws[target_cell].value = ws[source_cell].value
        #ws[target_cell]._style = copy(ws[source_cell]._style)
        ws[target_cell].font = copy(ws[source_cell].font)
        ws[target_cell].border = copy(ws[source_cell].border)
        ws[target_cell].fill = copy(ws[source_cell].fill)
        ws[target_cell].number_format = copy(ws[source_cell].number_format)
        ws[target_cell].protection = copy(ws[source_cell].protection)
        ws[target_cell].alignment = copy(ws[source_cell].alignment)




